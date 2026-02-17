from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from typing import List, Dict
import os
from config import EXCEL_DIR, EXCEL_COLUMNS


class ExcelWriter:
    def __init__(self):
        self.output_dir = EXCEL_DIR
    
    def write_jobs(self, jobs: List[Dict]) -> str:
        """Write jobs to Excel file with proper formatting"""
        # Create filename with date
        today = datetime.now().strftime("%d_%m_%Y")
        filename = f"Jobs_{today}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Job Applications"
        
        # Define styles
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col_num, header in enumerate(EXCEL_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Set column widths
        column_widths = {
            'A': 8,   # S.No
            'B': 12,  # Date
            'C': 15,  # Job ID
            'D': 20,  # Company
            'E': 25,  # Job Role
            'F': 40,  # Job Link
            'G': 15,  # Portal
            'H': 12,  # Status
            'I': 15,  # Interview Mode
            'J': 12,  # Interview Date
            'K': 12,  # Mail Sent
            'L': 12,  # Cold Email
            'M': 12,  # Follow Up
            'N': 12,  # Response
            'O': 30,  # Remarks
            'P': 30,  # Skills Required
            'Q': 10,  # Match %
            'R': 25,  # Resume Changes
            'S': 15   # Cover Letter
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Write data
        date_str = datetime.now().strftime("%d-%m-%Y")
        
        for idx, job in enumerate(jobs, 1):
            row = idx + 1
            
            # Auto-filled columns
            ws.cell(row=row, column=1, value=idx)  # S.No
            ws.cell(row=row, column=2, value=date_str)  # Date
            ws.cell(row=row, column=3, value=job.get('job_id', '-'))  # Job ID
            ws.cell(row=row, column=4, value=job.get('company', '-'))  # Company
            ws.cell(row=row, column=5, value=job.get('title', '-'))  # Job Role
            ws.cell(row=row, column=6, value=job.get('link', '-'))  # Job Link
            ws.cell(row=row, column=7, value=job.get('portal', '-'))  # Portal
            
            # Manual tracking fields (left blank)
            for col in range(8, 15):  # Status to Response Received
                ws.cell(row=row, column=col, value='')
            
            ws.cell(row=row, column=15, value='')  # Remarks
            
            # Skills and matching info
            skills_required = ', '.join(job.get('required_skills', []))
            ws.cell(row=row, column=16, value=skills_required)  # Exact Skills Required
            ws.cell(row=row, column=17, value=f"{job.get('match_percentage', 0):.1f}%")  # Match %
            
            missing_skills = ', '.join(job.get('missing_skills', []))
            ws.cell(row=row, column=18, value=missing_skills if missing_skills else 'None')  # Resume Changes
            
            ws.cell(row=row, column=19, value=job.get('cover_letter_generated', 'No'))  # Cover Letter
            
            # Apply styling to all cells in row
            for col in range(1, len(EXCEL_COLUMNS) + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = cell_alignment
                cell.border = border
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Save workbook
        wb.save(filepath)
        return filepath
